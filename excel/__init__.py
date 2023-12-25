# -*- coding = utf-8 -*-
# @Author : x_DARK_
# @File : __init__.py.py
# @Time : 2023/5/24 0024 15:51
# @Software : PyCharm

import sys

import openpyxl
import xlrd
import xlwings as xw

xlsx_in_path = r'F:\WorkSpace\pythonfile\excel\file\in.xlsx'
xlsx_out_path = r'F:\WorkSpace\pythonfile\excel\file\out.xlsx'

stu = []


def get_all():
    try:
        wb = openpyxl.load_workbook(xlsx_in_path)
        sheet = wb['all']
        cells = sheet['A']
        for cell in cells:
            stu.append({'name': cell.value, 'date': []})
        wb.close()
    except:
        print("get_all -> ?")
    #
    # stu[0]['姓名'].append('123')
    # print(stu[0]['姓名'])
    # ['123']
    # {'姓名': ['123']}
    # for v in stu:
    #     print(v)
    #
    pass


def get_sheet1():
    try:
        wb = openpyxl.load_workbook(xlsx_in_path)
        sheet = wb['sheet1']
        cells = sheet['A']
        for cell in cells:
            name = cell.value
            for i in range(len(stu)):
                if name == stu[i].get('name'):
                    stu[i].get('date').append('03.27')
                    break
                pass
        wb.close()
    except:
        print("get_sheet1 -> ?")
    #
    # for v in stu:
    #     print(v)
    pass


def get_sheet2():
    try:
        wb = openpyxl.load_workbook(xlsx_in_path)
        sheet = wb['sheet2']
        cells = sheet['A']
        for cell in cells:
            name = cell.value
            for i in range(len(stu)):
                if name == stu[i].get('name'):
                    stu[i].get('date').append('04.04')
                    break
                pass
        wb.close()
    except:
        print("get_sheet2 -> ?")
    #
    # count = 0
    # for v in stu:
    #     count += 1
    #     print(v)
    # print(count)
    pass


def get_sheet3():
    try:
        wb = openpyxl.load_workbook(xlsx_in_path)
        sheet = wb['sheet3']
        cells = sheet['A']
        for cell in cells:
            name = cell.value
            for i in range(len(stu)):
                if name == stu[i].get('name'):
                    stu[i].get('date').append('05.23')
                    break
                pass
        wb.close()
    except:
        print("get_sheet3 -> ?")
    #
    # count = 0
    # for v in stu:
    #     count += 1
    #     print(v)
    # print(count)
    pass


def out_excel():
    try:
        # 打开Excel程序，APP程序(即Excel程序)不可见，只打开不新建工作薄，屏幕更新关闭
        app = xw.App(visible=False, add_book=False)
        # Excel工作簿显示警告,不显示
        app.display_alerts = False
        # 工作簿屏幕更新,不更新
        app.screen_updating = False
        # 打开工作簿
        wb = app.books.open(xlsx_out_path)
        # 获取活动的工作表
        sheet = wb.sheets['Sheet1']

        # for i in range(0, len(stu), 1):
        #     print(stu[i].get('name'), end=' ')
        #     for s in stu[i].get('date'):
        #         print(s, end=' ')
        #     print()
        #     pass
        # sheet.range('A1').value = stu[0].get('name')
        for i in range(0, len(stu), 1):
            sheet.range('A' + str(i + 1)).value = stu[i].get('name')
            sheet.range('B' + str(i + 1) + ':' + 'C' + str(i + 1)).value = stu[i].get('date')
            print(stu[i].get('name'), end=' ')
            print(stu[i].get('date'))
            pass

        # 保存文件
        wb.save()
        # 关闭工作簿
        wb.close()
        # 退出Excel
        app.quit()
    except:
        print("output -> ?")
    pass


def main():
    get_all()
    get_sheet1()
    get_sheet2()
    get_sheet3()
    out_excel()


if __name__ == '__main__':
    main()
